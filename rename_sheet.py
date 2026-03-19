import openpyxl
import os
import sys

# Ensure stdout can handle Unicode characters
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def rename_first_sheet(file_path, new_name):
    """
    Renames the first sheet of an Excel file.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Get the first sheet (index 0)
        first_sheet_name = wb.sheetnames[0]
        sheet = wb[first_sheet_name]
        
        # Renaissance the sheet
        old_name = sheet.title
        sheet.title = new_name
        print(f"Successfully renamed sheet '{old_name}' to '{new_name}'")

        # Add a new worksheet if it doesn't exist
        instruction_sheet_name = "填寫說明"
        if instruction_sheet_name not in wb.sheetnames:
            wb.create_sheet(title=instruction_sheet_name)
            print(f"Successfully added new sheet '{instruction_sheet_name}'")
        else:
            print(f"Sheet '{instruction_sheet_name}' already exists.")
        
        # Save the workbook
        wb.save(file_path)
        print(f"Saved changes to {file_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")

def migrate_sheet_to_xlsm(src_path, dest_path, sheet_name):
    """
    Copies a specific sheet from a src_path XLSX to a dest_path XLSM, 
    preserving VBA macros in the destination.
    """
    if not os.path.exists(src_path):
        print(f"Error: Source file not found at {src_path}")
        return
    if not os.path.exists(dest_path):
        print(f"Error: Destination file not found at {dest_path}")
        return

    try:
        # Load source workbook
        wb_src = openpyxl.load_workbook(src_path)
        if sheet_name not in wb_src.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in source file.")
            return
        ws_src = wb_src[sheet_name]

        # Load destination XLSM (keep VBA macros)
        wb_dest = openpyxl.load_workbook(dest_path, keep_vba=True)
        
        # Get or create target sheet in destination
        if sheet_name in wb_dest.sheetnames:
            ws_dest = wb_dest[sheet_name]
            # Clear existing data
            for row in ws_dest.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws_dest = wb_dest.create_sheet(title=sheet_name)

        # Copy data row by row
        for row in ws_src.iter_rows():
            for cell in row:
                ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
        
        # Save destination XLSM (must end in .xlsm)
        wb_dest.save(dest_path)
        print(f"Successfully migrated sheet '{sheet_name}' to {dest_path}")

    except Exception as e:
        print(f"Migration error: {e}")

def delete_rows_from_sheet(file_path, sheet_name, start_row, count):
    """
    Deletes a range of rows from a specific sheet in an Excel file.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        # Load workbook (keep VBA for .xlsm)
        is_xlsm = file_path.lower().endswith(".xlsm")
        wb = openpyxl.load_workbook(file_path, keep_vba=is_xlsm)
        
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in {file_path}")
            return
        
        ws = wb[sheet_name]
        
        # Delete the rows
        ws.delete_rows(start_row, count)
        print(f"Successfully deleted {count} rows starting from {start_row} in sheet '{sheet_name}' of {file_path}")
        
        # Save changes
        wb.save(file_path)
        print(f"Saved changes to {file_path}")

    except Exception as e:
        print(f"Error deleting rows: {e}")

if __name__ == "__main__":
    # Original tasks
    target_xlsx = r"C:\Users\benga\Downloads\711_Batch_Import_03-19_00_41.xlsx"
    target_xlsm_main = r"C:\Users\benga\Desktop\Light-Local-MVP-main\賣貨便_訂單匯入.xlsm"
    
    # New row removal task
    target_xlsm_result = r"C:\Users\benga\Downloads\賣貨便_訂單匯入結果_2603191084664270.xlsm"
    
    sheet_import = "訂單匯入"
    
    # 1. XLSX processing
    rename_first_sheet(target_xlsx, sheet_import)
    
    # 2. Main migration
    migrate_sheet_to_xlsm(target_xlsx, target_xlsm_main, sheet_import)
    
    # 3. Specific row removal in result file
    delete_rows_from_sheet(target_xlsm_result, sheet_import, 9, 3)
