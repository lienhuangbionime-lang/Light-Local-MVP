import sys
import os

# [CRITICAL FIX] Render/Cloud Path Logic
# Always ensure the project root is in sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

print(f"[SYSTEM] Path Fix: current={current_dir}, parent={parent_dir}")

from fastapi import FastAPI, UploadFile, File, Form, Header, BackgroundTasks, HTTPException, Response, Request
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
import time
import json
import asyncio
import uuid
from typing import Dict, Optional, List
from fastapi.responses import StreamingResponse
import io
import xlsxwriter

# Modular Imports
from backend.config import (
    global_client, INSTANCE_ID, ORDER_POOL, LAST_EVENTS, PROCESSED_COMMENT_IDS,
    SESSION_START_TIME, IS_LIVE_ACTIVE, ADMIN_SECRET, CURRENT_PAGE_ID, ACTIVE_PRODUCTS
)
import backend.config as config
from backend.models.schemas import Order, ConfirmOrderRequest
from backend.core.security import verify_admin, verify_admin_signature, generate_admin_signature, verify_order_signature
from backend.database.firebase import (
    load_orders, load_events, save_orders, clear_orders_on_cloud, 
    sync_state_from_cloud, sync_orders_from_cloud, save_events
)
from backend.services.fb_service import (
    process_webhook_data, send_messenger_link, subscribe_page_to_app, save_fb_config, load_fb_config
)
from backend.services.order_service import process_order, handle_admin_secretarial_work
from backend.database.firebase import init_firebase

@asynccontextmanager
async def lifespan(app: FastAPI):
    """管理全域資源生命週期 (雲端加強版)"""
    print(f"🚀 [INIT] 啟動伺服器初始化... (Instance: {INSTANCE_ID})")
    
    try:
        # 1. Firebase 初始化
        init_firebase()
        
        # 2. 關鍵數據載入 (改為阻塞式確保啟動前準備完成)
        print("[INIT] 載入訂單與事件數據...")
        await load_orders()
        await load_events()
        
        # 3. 門市資料熱載入 (確保 store_service 已備妥)
        from backend.services.store_service import _load_stores_into_memory
        _load_stores_into_memory()

        # 4. FB 訂閱 (僅在有 Token 時執行，不佔用啟動時間)
        if config.PAGE_ACCESS_TOKEN:
            asyncio.create_task(subscribe_page_to_app(config.PAGE_ACCESS_TOKEN))
            
        print("✅ [INIT] 伺服器初始化成功")
    except Exception as e:
        print(f"❌ [INIT] 初始化崩潰: {e}")
        # 在雲端環境，初始化失敗通常代表程式無法運作，預留報錯
    
    yield
    
    # 關閉資源
    print("[EXIT] 正在關閉全域連線...")
    await global_client.aclose()

app = FastAPI(title="EchoOrder Buffer Gateway", lifespan=lifespan)

# CORS 配置：允許所有來源，但不攜帶憑證（因使用自定義 Header 驗證，需避開通配符 * 與 credentials=True 的衝突）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- 基本路由 ---

@app.get("/")
async def root():
    return {
        "message": "EchoOrder Buffer Gateway is running.",
        "docs": "/docs",
        "health": "/api/health",
        "instance": INSTANCE_ID
    }

@app.get("/api/health")
async def health_check():
    return {
        "status": "ok",
        "version": "2.1.0 (Modular Refactor)",
        "instance_id": INSTANCE_ID,
        "is_live": config.IS_LIVE_ACTIVE,
        "products_count": len(config.ACTIVE_PRODUCTS),
        "has_ai_key": bool(config.GEMINI_API_KEY)
    }

@app.get("/webhook/fb")
async def verify_fb_webhook(request: Request):
    params = request.query_params
    if params.get("hub.mode") == "subscribe" and params.get("hub.verify_token") == "echo_order_verify_token":
        return Response(content=params.get("hub.challenge"))
    return Response(content="Verification failed", status_code=403)

@app.post("/webhook/fb")
async def fb_webhook(request: Request, background_tasks: BackgroundTasks):
    try:
        data = await request.json()
        return await process_webhook_data(data, background_tasks)
    except Exception as e:
        print(f"[WEBHOOK] 解析失敗: {e}")
        return {"status": "error"}

# --- 賣家管理路由 ---

@app.get("/api/seller/orders/all")
async def get_all_orders():
    await sync_state_from_cloud(sync_orders=True)
    session_orders = [o.dict() for o in ORDER_POOL.values() if o.created_at >= config.SESSION_START_TIME]
    return {"orders": session_orders, "processed_comment_ids": list(PROCESSED_COMMENT_IDS)}

@app.post("/api/seller/sync_products")
async def update_products(request: Request, data: Dict):
    """更新商品代號表 (支援完整與巢狀結構)"""
    sig = request.headers.get("X-Admin-Signature")
    ts = request.headers.get("X-Admin-Timestamp")
    
    is_valid = verify_admin_signature(sig, ts)
    print(f"[DEBUG] Sync Products Attempt: ts={ts}, sig={sig}, valid={is_valid}")
    
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    # 自動拆解前端傳入的 { active_products: { ... }, free_shipping_threshold? } 
    products = data.get("active_products", {})
    threshold = data.get("free_shipping_threshold")
    
    if threshold is not None:
        config.FREE_SHIPPING_THRESHOLD = int(threshold)
    
    config.ACTIVE_PRODUCTS.clear()
    config.ACTIVE_PRODUCTS.update(products)
    await save_orders(save_config=True, fields=["active_products", "free_shipping_threshold"])
    return {"status": "success", "count": len(products), "threshold": config.FREE_SHIPPING_THRESHOLD}

# 舊端點相容別名
@app.post("/api/seller/active_products")
async def update_products_legacy(request: Request, data: Dict):
    return await update_products(request, data)

@app.post("/api/seller/config")
async def update_config(data: Dict):
    if "fb_page_token" in data:
        token = data["fb_page_token"].strip()
        config.PAGE_ACCESS_TOKEN = token
        save_fb_config({"fb_page_token": token})
    return {"status": "success"}

@app.post("/api/seller/subscribe_page")
async def subscribe_page(data: Dict):
    token = data.get("token") or config.PAGE_ACCESS_TOKEN
    return await subscribe_page_to_app(token)

@app.get("/api/seller/live/status")
async def get_live_status():
    return {"is_live_active": config.IS_LIVE_ACTIVE, "session_start_time": config.SESSION_START_TIME}

@app.post("/api/seller/live/status")
async def set_live_status(data: Dict):
    new_active = data.get("is_live_active", config.IS_LIVE_ACTIVE)
    
    # 🚀 [SESSION CLEANUP] 偵測結束直播 (從 True 變為 False)
    if config.IS_LIVE_ACTIVE and not new_active:
        print("[SYSTEM] 結束直播中... 正在清理與存檔訂單")
        pending_count = 0
        confirmed_count = 0
        
        # 1. 處理訂單池
        to_delete_local = []
        for oid, order in list(config.ORDER_POOL.items()):
            if order.status == "PENDING":
                # 未完成單 -> 直接刪除 (雲端也要清)
                to_delete_local.append(oid)
                pending_count += 1
            elif order.status == "CONFIRMED":
                # 已完成單 -> 搬移到 archived_orders 存檔 (Firestore 也搬)
                await save_orders(order_id=oid, move_to_archive=True)
                # 本地保留以便匯出，或者您可以選擇也從本地移除 (由 get_all_orders 的時間點決定)
                confirmed_count += 1
        
        # 清理本地 PENDING
        for oid in to_delete_local:
            if oid in config.ORDER_POOL:
                del config.ORDER_POOL[oid]
        
        # 2. 重置留言重複進單過濾器
        config.PROCESSED_COMMENT_IDS.clear()
        config.CURRENTLY_PROCESSING_IDS.clear()
        
        # 3. 呼叫雲端清理活躍訂單區 (確保 PENDING 鬼魂消失)
        await clear_orders_on_cloud(pending_only=True)
        
        print(f"[SYSTEM] 清理完成: 刪除 {pending_count} 筆未完成, 存檔 {confirmed_count} 筆已完成, 重置留言 ID")

    # 更新狀態
    config.IS_LIVE_ACTIVE = new_active
    if config.IS_LIVE_ACTIVE:
        # 開啟直播時重設起始時間
        config.SESSION_START_TIME = data.get("session_start_time", time.time())
    
    config.FREE_SHIPPING_THRESHOLD = data.get("free_shipping_threshold", config.FREE_SHIPPING_THRESHOLD)
    config.SHIPPING_FEE = data.get("shipping_fee", config.SHIPPING_FEE)
    
    # 保存配置
    await save_orders(save_config=True, fields=["is_live_active", "session_start_time", "free_shipping_threshold", "shipping_fee", "processed_comment_ids"])
    
    return {
        "status": "success", 
        "is_active": config.IS_LIVE_ACTIVE, 
        "session_start": config.SESSION_START_TIME
    }

@app.get("/api/seller/config")
async def get_seller_config(
    sig: str = Header(None, alias="X-Admin-Signature"),
    ts: str = Header(None, alias="X-Admin-Timestamp")
):
    verify_admin(ts, sig)
    return {
        "free_shipping_threshold": config.FREE_SHIPPING_THRESHOLD,
        "buyer_shipping_fee": config.BUYER_SHIPPING_FEE,
        "platform_shipping_fee": config.PLATFORM_SHIPPING_FEE
    }

@app.post("/api/seller/config")
async def update_seller_config(
    data: dict,
    sig: str = Header(None, alias="X-Admin-Signature"),
    ts: str = Header(None, alias="X-Admin-Timestamp")
):
    # print(f"[CONFIG] Received body: {data}") # DEBUG
    verify_admin(ts, sig)
    if "free_shipping_threshold" in data:
        config.FREE_SHIPPING_THRESHOLD = int(data["free_shipping_threshold"])
    if "buyer_shipping_fee" in data:
        config.BUYER_SHIPPING_FEE = int(data["buyer_shipping_fee"])
    if "platform_shipping_fee" in data:
        config.PLATFORM_SHIPPING_FEE = int(data["platform_shipping_fee"])
    
    # 🚀 [PERSISTENCE] Ensure changes are saved to cloud config
    await save_orders(save_config=True, fields=["free_shipping_threshold", "buyer_shipping_fee", "platform_shipping_fee"])
    
    return {"status": "ok"}

@app.post("/api/seller/sync_stores")
async def sync_stores(request: Request):
    """從 Firestore 同步 7-11 門市資料 (需要管理員簽名)"""
    sig = request.headers.get("X-Admin-Signature")
    ts = request.headers.get("X-Admin-Timestamp")
    if not verify_admin_signature(sig, ts):
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    from backend.database.firebase import sync_711_stores_from_cloud
    from backend.services.store_service import _load_stores_into_memory
    
    result = await sync_711_stores_from_cloud()
    if result["status"] == "success":
        # 同步後重新載入記憶體
        _load_stores_into_memory()
    return result

@app.get("/api/seller/stats")
async def get_stats():
    """統計銷售數據 (細分已給連結與已確認)"""
    stats = {}
    for order in ORDER_POOL.values():
        if order.created_at < config.SESSION_START_TIME: continue
        for item in order.items:
            code = item.product_code.upper()
            if code not in stats:
                stats[code] = {"pending": 0, "confirmed": 0}
            
            if order.status == "CONFIRMED" or order.status == "HARVESTED":
                stats[code]["confirmed"] += item.quantity
            else:
                stats[code]["pending"] += item.quantity
    return stats

@app.post("/api/seller/harvest")
async def run_harvest(request: Request):
    """批次將 PENDING 轉為 HARVESTED (需要管理員簽名)"""
    sig = request.headers.get("X-Admin-Signature")
    ts = request.headers.get("X-Admin-Timestamp")
    if not verify_admin_signature(sig, ts):
        print(f"[SECURITY] Harvest blocked: Invalid signature (sig={sig}, ts={ts})")
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    count = 0
    harvested_list = []
    # [FIX] 嚴格僅收割「買家已點過確認」的訂單 (CONFIRMED)
    for order in list(config.ORDER_POOL.values()):
        if order.status == "CONFIRMED":
            order.status = "HARVESTED"
            await save_orders(order_id=order.order_id)
            harvested_list.append(order.model_dump() if hasattr(order, 'model_dump') else order.dict())
            count += 1
    
    return {
        "status": "success", 
        "harvested": count, 
        "harvested_orders": harvested_list
    }

@app.delete("/api/seller/orders")
async def clear_orders_endpoint(request: Request, force: bool = False):
    sig = request.headers.get("X-Admin-Signature")
    ts = request.headers.get("X-Admin-Timestamp")
    
    is_valid = verify_admin_signature(sig, ts)
    if not is_valid:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    # 若不是 force (全清)，則只清除 PENDING (已給連結)
    # 若是 force，則清空全部
    pending_only = not force
    
    # 1. 清理本地 Cache
    if pending_only:
        # [REFINEMENT] 既然是要清空活躍區，則 ORDER_POOL 中所有「未搬移到 archived_orders」的訂單都該清除
        # 避免剛填完單但還沒導出的單子殘留在記憶體中，導致透過「自我修復」又被寫回雲端
        to_delete = [oid for oid, o in config.ORDER_POOL.items() if o.status != "HARVESTED"]
        for oid in to_delete:
            del config.ORDER_POOL[oid]
        # PROCESSED_COMMENT_IDS 不清理，避免重覆進單 (除非全清)
    else:
        # 全清
        config.ORDER_POOL.clear()
        config.PROCESSED_COMMENT_IDS.clear()
    
    # 2. 清理雲端
    deleted = await clear_orders_on_cloud(pending_only=pending_only)
    return {"status": "success", "count": deleted, "mode": "pending_only" if pending_only else "all"}

@app.get("/api/debug/time")
async def get_server_time():
    """檢查伺服器時間與 Secret 狀態 (Debug 用)"""
    return {
        "server_time": time.time(),
        "admin_secret_set": len(config.ADMIN_SECRET) > 0,
        "admin_secret_preview": f"{config.ADMIN_SECRET[:3]}***" if config.ADMIN_SECRET else "NONE"
    }

# --- 買家結帳路由 ---

@app.get("/api/checkout/{order_id}")
async def get_checkout(order_id: str, s: Optional[str] = None):
    # 如果有帶 s，也要驗證一下安全性 (雖然原代碼只有 GET，但補上驗證更好)
    if s and not verify_order_signature(order_id, s):
        raise HTTPException(status_code=403, detail="Invalid Signature")
        
    if order_id not in ORDER_POOL: await sync_orders_from_cloud()
    # 為了計算方便，將最新的 price_rule 注入到訂單項目的 metadata 中
    if order_id in ORDER_POOL:
        order_data = ORDER_POOL[order_id].model_dump()
        # 注入全域免運門檻與商品中繼資料
        order_data["config"] = {
            "products": config.ACTIVE_PRODUCTS,
            "free_shipping_threshold": config.FREE_SHIPPING_THRESHOLD,
            "shipping_fee": config.SHIPPING_FEE,
            "buyer_shipping_fee": config.BUYER_SHIPPING_FEE,
            "platform_shipping_fee": config.PLATFORM_SHIPPING_FEE
        }
        for item in order_data.get("items", []):
            p_data = config.ACTIVE_PRODUCTS.get(item["product_code"], {})
            if isinstance(p_data, dict):
                item["price_rule"] = p_data.get("price_rule", "")
        return order_data
    raise HTTPException(status_code=404, detail="Order not found")

async def do_ai_fill_task(order_id: str, image_b64: str):
    """🧠 Background Task for AI Vision processing"""
    try:
        if order_id not in config.ORDER_POOL:
            await sync_orders_from_cloud()
        if order_id not in config.ORDER_POOL:
            return

        order = config.ORDER_POOL[order_id]
        order.ai_status = "processing"
        from backend.database.firebase import save_orders
        await save_orders(order_id=order_id, fields=["ai_status"])

        prompt = f"""你現在是【Role: EchoOrder Checkout Helper】。
你的任務是從買家提供的「截圖」中，精確提取收件資訊。

請【嚴格】回傳以下 JSON 格式。

【規則 (最高優先級)】：
1. **多重探測 (Candidates)**：在 `store_candidates` 中列出所有你在圖中看到的「店名」或「6位數字」。
2. **數字排除**：
   - 看到年份數字 (如 113, 2024, 2025)：**禁止放入 candidates**。
   - 看到發票號碼 (含英文或 8 位以上純數字)：**禁止放入 candidates**。
3. **店名提取**：請提取如「旗山旗力」、「港富」等獨特名字。
4. **買家姓名**：優先抓取對話視窗最頂部顯示的對象名稱。

【JSON 格式要求】：
{{
  "buyer_name": "買家姓名",
  "phone": "10 碼電話",
  "store_candidates": ["店號1", "店名1", "店號2"...]
}}"""
        
        extracted = await ask_gemini_secretary(text_content="[USER PHOTO FILL]", image_data_base64=image_b64, system_prompt=prompt)
        
        if not extracted or not isinstance(extracted, dict):
            extracted = {"buyer_name": "", "phone": "", "store_candidates": [], "items": []}
        
        if extracted.get("phone"):
            extracted["phone"] = "".join(filter(str.isdigit, str(extracted["phone"])))
        
        from backend.services.ai_service import transcribe_image_text
        from backend.services.parse_service import extract_store_candidates_from_ocr, extract_store_names_from_ocr
        from backend.services.store_service import resolve_store_info, _STORE_BY_NAME
        
        ocr_text = await transcribe_image_text(image_b64)
        ocr_candidates = []
        if ocr_text:
            ocr_candidates = extract_store_candidates_from_ocr(ocr_text)
            name_matched_ids = extract_store_names_from_ocr(ocr_text, _STORE_BY_NAME)
            ocr_candidates = name_matched_ids + ocr_candidates
        
        ai_candidates = extracted.get("store_candidates", [])
        all_candidates = ocr_candidates + ai_candidates
        
        verified_store = await resolve_store_info("", candidates=all_candidates)
        extracted["shipping_info"] = verified_store
        
        # Apply to order
        has_update = False
        if extracted.get("buyer_name") and not order.buyer_name:
            order.buyer_name = extracted["buyer_name"]; has_update = True
        if extracted.get("phone") and not order.phone:
            order.phone = extracted["phone"]; has_update = True
        if extracted.get("shipping_info") and not order.shipping_info:
            order.shipping_info = extracted["shipping_info"]; has_update = True
        
        order.ai_status = "done"
        order.ai_error = None
        await save_orders(order_id=order_id)
        print(f"[AI_FILL] Order {order_id} processed successfully.")

    except Exception as e:
        print(f"[AI_FILL] Background task failed: {e}")
        if order_id in config.ORDER_POOL:
            config.ORDER_POOL[order_id].ai_status = "failed"
            config.ORDER_POOL[order_id].ai_error = str(e)
            from backend.database.firebase import save_orders
            await save_orders(order_id=order_id, fields=["ai_status", "ai_error"])

@app.post("/api/checkout/{order_id}/ai_fill")
async def ai_fill_order(order_id: str, data: Dict, background_tasks: BackgroundTasks, s: Optional[str] = None):
    if not verify_order_signature(order_id, s):
        raise HTTPException(status_code=403, detail="Invalid Signature")
    
    image_b64 = data.get("image")
    if not image_b64:
        return {"status": "error", "message": "Missing image data"}

    if order_id not in config.ORDER_POOL:
        await sync_orders_from_cloud()
    
    if order_id in config.ORDER_POOL:
        order = config.ORDER_POOL[order_id]
        if order.ai_status == "processing":
            return {"status": "processing", "message": "Already processing"}
        
        order.ai_status = "processing"
        background_tasks.add_task(do_ai_fill_task, order_id, image_b64)
        return {"status": "processing", "message": "Job started in background"}

    return {"status": "error", "message": "Order not found"}

@app.get("/api/checkout/{order_id}/status")
async def get_order_ai_status(order_id: str, s: Optional[str] = None):
    if not verify_order_signature(order_id, s):
        raise HTTPException(status_code=403, detail="Invalid Signature")
    
    if order_id not in config.ORDER_POOL:
        await sync_orders_from_cloud()
        
    if order_id in config.ORDER_POOL:
        order = config.ORDER_POOL[order_id]
        return {
            "order_id": order_id,
            "ai_status": order.ai_status,
            "ai_error": order.ai_error,
            "data": {
                "buyer_name": order.buyer_name,
                "phone": order.phone,
                "shipping_info": order.shipping_info
            }
        }
    raise HTTPException(status_code=404, detail="Order not found")

@app.post("/api/checkout/{order_id}/confirm")
async def confirm_checkout(order_id: str, data: Dict, s: Optional[str] = None):
    if not verify_order_signature(order_id, s):
        raise HTTPException(status_code=403, detail="Invalid Signature")
    if order_id not in ORDER_POOL: await sync_orders_from_cloud()
    if order_id in ORDER_POOL:
        order = ORDER_POOL[order_id]
        order.buyer_name = data.get("buyer_name")
        order.shipping_info = data.get("shipping_info")
        order.phone = data.get("phone")
        order.status = "CONFIRMED"
        # 立即搬移到封存區 (archived_orders)，讓主 orders 保持清空
        await save_orders(order_id=order_id, move_to_archive=True)
        return {"status": "success"}
    raise HTTPException(status_code=404, detail="Order not found")

# --- Debug & Utility ---

@app.get("/api/debug/events")
async def get_debug_events():
    return config.LAST_EVENTS

@app.get("/api/admin/config_check")
async def check_config(admin_secret: Optional[str] = None):
    # 此處可供使用者在線上檢查環境變數是否設定正確
    if admin_secret != config.ADMIN_SECRET:
        return {"status": "error", "message": "Unauthorized"}
    
    return {
        "GEMINI_API_KEY": f"SET (ends with {config.GEMINI_API_KEY[-4:]})" if config.GEMINI_API_KEY else "MISSING ❌",
        "PAGE_ACCESS_TOKEN": "SET" if config.PAGE_ACCESS_TOKEN else "MISSING ❌",
        "INFO": "請確保在 Render.com 後台的 Environment Variables 正確設定了這些值。"
    }

@app.get("/api/debug/products")
async def get_debug_products():
    return {
        "product_count": len(config.ACTIVE_PRODUCTS),
        "active_products": config.ACTIVE_PRODUCTS
    }

@app.post("/api/seller/orders/restore")
async def restore_orders(request: Request):
    """前端定期備份/自我修復：將失去的訂單或留言 ID 同步回後端"""
    data = await request.json()
    restored_count = 0
    orders_data = data.get("orders", [])
    
    for o_dict in orders_data:
        order_id = o_dict.get("order_id")
        if order_id and order_id not in config.ORDER_POOL:
            # 直接把前端字典塞進 Pydantic Model，這會完美保留 json 裡的 created_at 時間戳！
            order = Order(**o_dict)
            config.ORDER_POOL[order_id] = order
            restored_count += 1
            
    processed_ids = data.get("processed_comment_ids", [])
    for pid in processed_ids:
        if pid not in config.PROCESSED_COMMENT_IDS:
            config.PROCESSED_COMMENT_IDS.add(pid)
            
    if restored_count > 0 or processed_ids:
        await save_orders()
        print(f"[SELF-HEALING] 恢復了 {restored_count} 筆訂單與 {len(processed_ids)} 筆留言紀錄")
        
    return {"status": "success", "restored": restored_count}

@app.post("/api/seller/pull_live_comments")
async def pull_live_comments_stub(request: Request):
    """主動抓取直播留言的端點 (目前僅為靜音 404 報錯)"""
    return {"status": "success", "new_orders": 0, "message": "Stub: Polling not fully implemented"}

@app.delete("/api/debug/events")
async def clear_events():
    print(f"[DEBUG] 清除事件紀錄前：{len(config.LAST_EVENTS)} 筆")
    config.LAST_EVENTS.clear()
    from backend.database.firebase import save_events
    await save_events()
    print(f"[DEBUG] 清除成功")
    return {"status": "success"}

@app.post("/api/debug/simulate_webhook")
async def simulate_webhook(data: Dict, background_tasks: BackgroundTasks):
    """
    開發用：允許兩種 payload
    1) 完整 FB Webhook（含 entry/changes/messaging）
    2) 簡化測試格式：{ code: "A", quantity: 2, sender_id?, sender_name? }
    """
    if "entry" not in data and "code" in data:
        code = str(data.get("code", "")).strip().upper()
        qty = int(data.get("quantity", 1) or 1)
        sender_id = str(data.get("sender_id", "u_sim"))
        sender_name = str(data.get("sender_name", "Simulator"))
        # Wrap into FB feed comment change format expected by process_webhook_data
        data = {
            "object": "page",
            "entry": [
                {
                    "id": config.CURRENT_PAGE_ID or "page_sim",
                    "time": int(time.time()),
                    "changes": [
                        {
                            "field": "feed",
                            "value": {
                                "item": "comment",
                                "message": f"{code}+{qty}",
                                "from": {"id": sender_id, "name": sender_name},
                                "comment_id": f"sim_{uuid.uuid4().hex[:8]}",
                            },
                        }
                    ],
                }
            ],
        }

    return await process_webhook_data(data, background_tasks, is_simulated=True)

@app.post("/api/admin/emergency/recover_order")
async def recover_order(data: Dict):
    """
    緊急補單功能：手動恢復遺失的 PENDING 訂單
    payload: { admin_secret, order_id, buyer_name, items: [{code, qty}] }
    """
    if data.get("admin_secret") != config.ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    order_id = data.get("order_id")
    if not order_id: return {"status": "error", "message": "Missing order_id"}
    
    order_items = []
    from backend.models.schemas import OrderItem
    for it in data.get("items", []):
        code = it.get("code", "").upper()
        qty = it.get("qty", 1)
        p_data = config.ACTIVE_PRODUCTS.get(code, {})
        p_name = p_data.get("name", code) if isinstance(p_data, dict) else p_data
        p_rule = p_data.get("price_rule", "") if isinstance(p_data, dict) else ""
        from backend.services.order_service import get_price_from_rule
        total_p = get_price_from_rule(p_rule, qty)
        unit_p = total_p // qty if qty > 0 else 0
        order_items.append(OrderItem(product_code=code, quantity=qty, product_name=p_name, price=unit_p))

    new_order = Order(
        order_id=order_id,
        fb_user_name=data.get("buyer_name", "Recovered User"),
        items=order_items,
        status="PENDING",
        created_at=time.time(),
        instance_id=config.INSTANCE_ID
    )
    
    config.ORDER_POOL[order_id] = new_order
    await save_orders(order_id=order_id)
    return {"status": "success", "message": f"Order {order_id} recovered to database"}

@app.get("/api/seller/orders/export_xlsx")
async def export_orders():
    """
    導出符合 7-11 賣貨便 (MyShip) 批量進貨格式的 Excel
    格式：收件人姓名, 收件人手機, 取貨門市店號, 溫層, 商品名稱, 訂單金額(代收), 運費金額
    """
    await sync_state_from_cloud(sync_orders=True)

    # [TEMPLATE] 使用用戶提供的賣貨便官方範本
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, "assets", "template.xlsm")
    
    if not os.path.exists(template_path):
        print(f"[EXPORT] 找不到範本檔案: {template_path}")
        return {"status": "error", "message": "伺服器遺失 Excel 範本檔案 (template.xlsm)"}

    output = io.BytesIO()
    try:
        import openpyxl
        # 載入範本，但不保留 VBA 巨集以加速匯出並產生標準 .xlsx
        wb = openpyxl.load_workbook(template_path)
        
        # 確保分頁名稱正確
        if "訂單匯入" not in wb.sheetnames:
            print(f"[EXPORT] 範本格式錯誤，找不到 '訂單匯入' 分頁. 現有分頁: {wb.sheetnames}")
            return {"status": "error", "message": "Excel 範本格式不符 (缺少 '訂單匯入' 分頁)"}
            
        ws = wb["訂單匯入"]
        
        # 🚀 [CONCURRENCY SAFE] 建立唯讀快照
        snapshot_orders = list(config.ORDER_POOL.values())
        merged_orders = {}
        for o in snapshot_orders:
            # 🚀 嚴格僅導出「已確認 (CONFIRMED)」的訂單
            if o.status != "CONFIRMED":
                continue
            
            b_name = (o.buyer_name or o.fb_user_name or "未命名").strip()
            phone = (o.phone or "").strip()
            store = (o.shipping_info or "").strip()
            key = f"{phone}_{store}_{b_name}"
            
            if key not in merged_orders:
                merged_orders[key] = {
                    "buyer_name": b_name,
                    "phone": phone,
                    "shipping_info": store,
                    "items": []
                }
            merged_orders[key]["items"].extend(o.items)

        current_row = 7 # 根據用戶範本，資料從第 7 列開始
        platform_shipping = config.PLATFORM_SHIPPING_FEE
        
        for key, data in merged_orders.items():
            items_price = sum((item.price or 0) * item.quantity for item in data["items"])
            total_qty = sum(item.quantity for item in data["items"])
            
            is_free = total_qty >= config.FREE_SHIPPING_THRESHOLD
            total_shipping_charged = 0 if is_free else config.BUYER_SHIPPING_FEE
            total_amount = int(round(items_price + total_shipping_charged))
            
            # 商品彙整
            from collections import Counter
            counts = Counter()
            for i in data["items"]:
                counts[i.product_code or "未知"] += i.quantity
            items_summary = ", ".join([f"{code}x{qty}" for code, qty in counts.items()])
            
            # 提取 6 碼門市店號
            import re
            m = re.search(r'\d{6}', data["shipping_info"] or "")
            store_id = m.group(0) if m else data["shipping_info"]
            
            # 寫入 Excel (對齊範本 12 欄位，但只填寫前 7 欄)
            ws.cell(row=current_row, column=1, value=data["buyer_name"])
            ws.cell(row=current_row, column=2, value=data["phone"])
            ws.cell(row=current_row, column=3, value=store_id)
            ws.cell(row=current_row, column=4, value="常溫")
            ws.cell(row=current_row, column=5, value=items_summary)
            # 訂單金額 = 總額 - 平台運費
            ws.cell(row=current_row, column=6, value=max(0, total_amount - platform_shipping))
            ws.cell(row=current_row, column=7, value=platform_shipping)
            current_row += 1
            
        wb.save(output)
        output.seek(0)
        
        filename = f"711_Batch_Import_{time.strftime('%m-%d_%H_%M')}.xlsx"
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"[EXPORT] Excel 導出失敗: {e}")
        return {"status": "error", "message": f"Excel 處理失敗: {str(e)}"}

if __name__ == "__main__":
    import uvicorn
    # Render 會注入 PORT 環境變數，若無則預設 8000
    port_str = os.environ.get("PORT", "8000")
    port = int(port_str)
    print(f"🔥 [SYSTEM] 啟動伺服器於 Port: {port}")
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
else:
    # 這是被 uvicorn 命令列啟動時 (例如 Cloud/Docker)
    import os
    port = os.environ.get("PORT", "Unknown")
    print(f"📡 [SYSTEM] 模組載入中... (RUNNING_ON_CLOUD, PORT={port})")
